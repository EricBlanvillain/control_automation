import os
import logging
import io # Needed for OCR image processing if used
from typing import List, Tuple, Optional, Dict
from agno.agent import Agent
# Import necessary libraries for file reading
import docx
import openpyxl
# For PDF/OCR, integration is needed. Placeholder libraries:
# import fitz  # No longer needed for PDF text extraction
import uuid

# -- Added/Modified Imports ---
import chromadb
from chromadb.utils import embedding_functions
from openai import OpenAI, OpenAIError # Keep for embeddings
import mistralai # Added Mistral AI client
from mistralai import Mistral # Correct import for the client
# -- Removed OCR Imports ---
# import pytesseract # Removed
# from PIL import Image # Removed
# --- End Added/Modified Imports ---
# --- Import shared utility ---
from utils.embedding_utils import get_openai_embeddings
# --- End Import ---

# Set up logger
logger = logging.getLogger(__name__)

# TODO: Research and implement Mistral OCR integration if required.
#       This might involve using a specific API, library, or a custom Agno tool.

# Configuration for chunking (can be overridden by env vars)
DEFAULT_CHUNK_SIZE = int(os.getenv("EXTRACTOR_CHUNK_SIZE", 2000))
DEFAULT_CHUNK_OVERLAP = int(os.getenv("EXTRACTOR_CHUNK_OVERLAP", 200))
CHUNK_SIZE = DEFAULT_CHUNK_SIZE
CHUNK_OVERLAP = DEFAULT_CHUNK_OVERLAP

# OpenAI Configuration
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
# Mistral Configuration
MISTRAL_API_KEY = os.getenv("MISTRAL_API_KEY")

logger.info(f"Extractor configured with Chunk Size: {CHUNK_SIZE}, Overlap: {CHUNK_OVERLAP}")
if not OPENAI_API_KEY:
    logger.warning("OPENAI_API_KEY environment variable not set. OpenAI Embeddings will fail.")
if not MISTRAL_API_KEY:
    logger.warning("MISTRAL_API_KEY environment variable not set. Mistral OCR will fail.")

# Define return type for extraction methods: (content: str | None, error_msg: str | None)
ExtractionResult = Tuple[Optional[str], Optional[str]]
# Corrected return type for the main run method
RunResult = Tuple[Optional[chromadb.Collection], Optional[str]]

class ExtractorAgent:
    """
    Agent 2: Extracts text from various document types (using Mistral OCR for PDFs),
    chunks it, generates embeddings, and stores them in an in-memory vector store.
    """
    def __init__(self):
        logger.info("Initializing ExtractorAgent...")
        self.openai_client = None
        self.mistral_client = None
        self.chroma_client = None
        try:
            logger.critical("ENTERED ExtractorAgent __init__ TRY block")

            # Initialize OpenAI client (for embeddings)
            logger.debug("Attempting OpenAI client init...")
            if OPENAI_API_KEY:
                self.openai_client = OpenAI(api_key=OPENAI_API_KEY)
                logger.debug("OpenAI client potentially initialized.")
            else:
                logger.error("OpenAI client skipped: API key missing.")

            # Initialize Mistral client (for OCR)
            logger.debug("Attempting Mistral client init...")
            if MISTRAL_API_KEY:
                self.mistral_client = Mistral(api_key=MISTRAL_API_KEY)
                logger.debug("Mistral client potentially initialized.")
            else:
                logger.error("Mistral client skipped: API key missing.")

            # Initialize ChromaDB client
            logger.debug("Attempting ChromaDB client init...")
            data_path = os.path.join(os.getcwd(), ".chroma_data")
            logger.debug(f"ChromaDB data path: {data_path}")
            self.chroma_client = chromadb.PersistentClient(path=data_path)
            logger.debug("ChromaDB PersistentClient call completed.")
            if self.chroma_client:
                logger.debug(f"ChromaDB client object CHECK successful (type: {type(self.chroma_client)})")
            else:
                logger.error(f"ChromaDB client object CHECK FAILED (client is None/Falsy) after init call.")
            logger.info(f"ChromaDB client initialization appears complete. Path: {data_path}")

            logger.critical("EXITED ExtractorAgent __init__ TRY block (Successful)")

        except BaseException as e:
            logger.critical("ENTERING ExtractorAgent __init__ EXCEPT block")
            logger.error(f"!!! EXCEPTION during ExtractorAgent component initialization: {e} !!!", exc_info=True)

        # Check client status immediately after try-except
        if not self.chroma_client:
            logger.warning("ExtractorAgent __init__ Check: self.chroma_client is None AFTER try-except.")
        if not self.openai_client:
            logger.warning("ExtractorAgent __init__ Check: self.openai_client is None AFTER try-except.")

        logger.info("ExtractorAgent initialized.")

    def _extract_from_docx(self, file_path: str) -> ExtractionResult:
        """Extracts text from a .docx file."""
        logger.debug(f"Extracting text from DOCX: {file_path}")
        try:
            doc = docx.Document(file_path)
            full_text = [para.text for para in doc.paragraphs if para.text]
            content = "\n".join(full_text)
            return content, None
        except Exception as e:
            error_msg = f"Error reading DOCX {file_path}: {e}"
            logger.error(error_msg, exc_info=True)
            return None, error_msg

    def _extract_from_xlsx(self, file_path: str) -> ExtractionResult:
        """Extracts text from an .xlsx file."""
        logger.debug(f"Extracting text from XLSX: {file_path}")
        workbook = None
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            full_text = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                logger.debug(f"--- Reading Sheet: {sheet_name} ---")
                for row in sheet.iter_rows():
                    row_text = [str(cell.value) for cell in row if cell.value is not None]
                    if row_text:
                        full_text.append(" | ".join(row_text))
            content = "\n".join(full_text)
            return content, None
        except Exception as e:
            error_msg = f"Error reading XLSX {file_path}: {e}"
            logger.error(error_msg, exc_info=True)
            return None, error_msg
        finally:
             if workbook:
                 try: workbook.close()
                 except: pass # Ignore errors on close
                 logger.debug(f"Closed XLSX workbook: {file_path}")

    def _extract_from_pdf(self, file_path: str) -> ExtractionResult:
        """Extracts text from a .pdf file using the Mistral OCR API."""
        logger.info(f"Attempting PDF extraction for {file_path} using Mistral OCR...")
        if not self.mistral_client:
            error_msg = "Error: Mistral client not initialized. Cannot perform OCR."
            logger.error(error_msg)
            return None, error_msg

        uploaded_file_id = None
        try:
            # 1. Upload the local file to Mistral - CORRECTED METHOD NAME and ARGUMENT FORMAT
            logger.debug(f"Uploading file {file_path} to Mistral for OCR...")
            with open(file_path, "rb") as f:
                # Pass file argument as a dictionary matching Mistral SDK expectation
                uploaded_pdf = self.mistral_client.files.upload(
                    file={
                        "file_name": os.path.basename(file_path),
                        "content": f
                    },
                    purpose="ocr"
                )
            uploaded_file_id = uploaded_pdf.id
            logger.debug(f"File uploaded successfully. File ID: {uploaded_file_id}")

            # 2. Get Signed URL for the uploaded file
            logger.debug(f"Getting signed URL for file ID: {uploaded_file_id}")
            signed_url = self.mistral_client.files.get_signed_url(file_id=uploaded_file_id)
            logger.debug(f"Got signed URL: {signed_url.url[:50]}...")

            # 3. Process the file using the signed URL with the OCR model
            logger.info(f"Processing file via signed URL ({signed_url.url[:50]}...) with Mistral OCR model...")
            ocr_response = self.mistral_client.ocr.process(
                model="mistral-ocr-latest",
                # Use the document URL argument type
                document={
                    "type": "document_url",
                    "document_url": signed_url.url
                }
            )
            logger.debug(f"Mistral OCR processing complete for file ID: {uploaded_file_id}")

            # 4. Concatenate markdown content from all pages
            full_markdown_content = "\n".join([page.markdown for page in ocr_response.pages])

            # 5. Cleanup uploaded file (always attempt)
            try:
                logger.debug(f"Deleting uploaded Mistral file: {uploaded_file_id}")
                self.mistral_client.files.delete(file_id=uploaded_file_id)
            except Exception as del_e:
                logger.warning(f"Failed to delete uploaded Mistral file {uploaded_file_id}: {del_e}")

            # 6. Return content if found
            if full_markdown_content.strip():
                logger.info(f"Successfully extracted markdown content via Mistral OCR from {file_path}. Total length: {len(full_markdown_content)}")
                return full_markdown_content, None
            else:
                error_msg = "Error: Mistral OCR processed the file but returned no content."
                logger.error(error_msg + f" for {file_path} (File ID: {uploaded_file_id})")
                return None, error_msg

        except Exception as e:
            # Check if it *might* be an API error based on type/string - enhance logging
            is_mistral_api_error = hasattr(e, '__class__') and 'Mistral' in e.__class__.__name__
            error_prefix = "Mistral API error" if is_mistral_api_error else "Unexpected error"
            error_msg = f"Error: {error_prefix} during PDF processing: {e}"
            logger.error(f"{error_msg} for {file_path}", exc_info=True)
            # Attempt cleanup if upload succeeded but process failed
            if uploaded_file_id:
                try: self.mistral_client.files.delete(file_id=uploaded_file_id)
                except Exception: pass
            return None, error_msg
        except FileNotFoundError:
            error_msg = f"Error: Local file not found: {file_path}"
            logger.error(error_msg)
            return None, error_msg
        except Exception as e:
            error_msg = f"Error: Unexpected error during Mistral PDF processing ({e})"
            logger.error(f"Error processing PDF {file_path}: {e}", exc_info=True)
            # Attempt cleanup if upload succeeded but process failed
            if uploaded_file_id:
                try: self.mistral_client.files.delete(file_id=uploaded_file_id)
                except Exception: pass
            return None, error_msg

    def _extract_from_txt(self, file_path: str) -> ExtractionResult:
        """Extracts text from a plain .txt file."""
        logger.debug(f"Extracting text from TXT: {file_path}")
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            return content, None
        except Exception as e:
            error_msg = f"Error reading TXT {file_path}: {e}"
            logger.error(error_msg, exc_info=True)
            return None, error_msg

    def _chunk_text(self, text: str, chunk_size: int, chunk_overlap: int) -> List[str]:
        """Splits text into chunks of a specified size with overlap."""
        # TODO: Current chunking is naive for Markdown. Consider parsing Markdown
        #       and chunking based on semantic structure (headers, paragraphs) for
        #       better context preservation in chunks.
        if chunk_size <= 0:
             logger.error(f"Chunk size ({chunk_size}) must be positive.")
             return [text]
        if chunk_overlap >= chunk_size:
             logger.warning(f"Overlap ({chunk_overlap}) >= size ({chunk_size}). Setting overlap to {max(0, chunk_size // 2)}.\n")
             chunk_overlap = max(0, chunk_size // 2)
        elif chunk_overlap < 0:
             logger.warning(f"Overlap ({chunk_overlap}) is negative. Setting to 0.\n")
             chunk_overlap = 0

        if not text:
            logger.warning("Received empty text for chunking.")
            return []

        logger.debug(f"Chunking text (length {len(text)}) with size {chunk_size}, overlap {chunk_overlap}")
        chunks = []
        start = 0
        while start < len(text):
            end = start + chunk_size
            chunks.append(text[start:end])
            next_start = start + chunk_size - chunk_overlap
            if next_start <= start:
                break
            start = next_start
        logger.info(f"Split text into {len(chunks)} chunks.")
        return chunks

    def run(self, file_path: str) -> RunResult:
        """
        Extracts text, chunks it, generates embeddings, stores in ChromaDB.
        Returns: Tuple (Optional[chromadb.Collection], Optional[str]) # Corrected return type hint
                 - (chroma_collection, None) on success.
                 - (None, error_message) on failure.
        """
        logger.info(f"Extractor starting processing for document: {file_path}")
        if not os.path.exists(file_path):
            error_msg = f"Error: File not found: {file_path}"
            logger.error(error_msg)
            return None, error_msg
        if not self.openai_client:
            error_msg = "Error: ExtractorAgent not fully initialized (OpenAI client missing). Check OPENAI_API_KEY."
            logger.error(error_msg)
            return None, error_msg
        if not self.chroma_client:
            error_msg = "Error: ExtractorAgent not fully initialized (Chroma client missing). Check ChromaDB setup/permissions."
            logger.error(error_msg)
            return None, error_msg

        _, extension = os.path.splitext(file_path.lower())
        content: Optional[str] = None
        error_msg: Optional[str] = None

        logger.info(f"Detected file extension '{extension}'. Routing to appropriate extractor.")
        extractor_method = {
            '.docx': self._extract_from_docx,
            '.xlsx': self._extract_from_xlsx,
            '.pdf': self._extract_from_pdf,
            '.txt': self._extract_from_txt
        }.get(extension)

        if extractor_method:
            content, error_msg = extractor_method(file_path)
        else:
            error_msg = f"Error: Unsupported file type: {extension}"
            logger.error(error_msg)
            return None, error_msg

        # --- Handle Extraction Failure ---
        if error_msg:
             logger.error(f"Extraction failed for {file_path}: {error_msg}")
             return None, error_msg
        if content is None or content == "":
             logger.warning(f"Extraction yielded no text content for {file_path}. Cannot proceed.")
             # Return success but with no collection, indicating no content to process
             return None, None # Or perhaps return an empty collection marker?

        # --- Chunking ---
        logger.info(f"Finished extraction for {file_path}. Extracted text length: {len(content)}. Proceeding to chunking.")
        text_chunks = self._chunk_text(content, chunk_size=CHUNK_SIZE, chunk_overlap=CHUNK_OVERLAP)

        if not text_chunks:
            logger.warning(f"Chunking resulted in zero chunks for {file_path}. Cannot proceed.")
            return None, None # Success, but no chunks to embed

        # --- Embedding --- NOW THIS CODE WILL EXECUTE ---
        embeddings = get_openai_embeddings(self.openai_client, text_chunks)
        if embeddings is None:
            # Error is already logged by the utility function
            return None, "Error: Failed to generate embeddings."

        if len(embeddings) != len(text_chunks):
             error_msg = f"Error: Mismatch between number of chunks ({len(text_chunks)}) and embeddings ({len(embeddings)})."
             logger.error(error_msg)
             return None, error_msg

        # --- Vector Store Population --- NOW THIS CODE WILL EXECUTE ---
        collection_name = f"doc_{uuid.uuid4().hex}" # Define here for cleanup scope
        try:
            logger.info(f"Creating/getting ChromaDB collection: {collection_name}")

            # Get or create the collection.
            collection = self.chroma_client.get_or_create_collection(
                name=collection_name,
                 metadata={"hnsw:space": "cosine"} # Use cosine distance
            )

            # Prepare data for ChromaDB: IDs, embeddings, documents (chunks)
            chunk_ids = [f"chunk_{i}" for i in range(len(text_chunks))]

            logger.info(f"Adding {len(text_chunks)} chunks with embeddings to collection '{collection_name}'...")
            # Add in batches if necessary for very large documents
            collection.add(
                embeddings=embeddings,
                documents=text_chunks,
                ids=chunk_ids
            )
            logger.info(f"Successfully added data to ChromaDB collection '{collection_name}'.")

            # Return the populated collection - THIS IS THE CORRECT RETURN
            return collection, None

        except Exception as e:
            error_msg = f"Error interacting with ChromaDB: {e}"
            logger.error(error_msg, exc_info=True)
            # Attempt cleanup
            try:
                self.chroma_client.delete_collection(collection_name)
                logger.info(f"Attempted cleanup: Deleted potentially partial collection '{collection_name}'.")
            except Exception as cleanup_e:
                 logger.error(f"Error during ChromaDB cleanup for collection '{collection_name}': {cleanup_e}")
            return None, error_msg
